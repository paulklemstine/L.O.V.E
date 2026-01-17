"""
AbundanceGenerator Module - Automated Resource Acquisition System

A modular system for automating and optimizing resource acquisition through
algorithmic trading, content monetization, and digital asset management.

Features:
- Algorithmic trading using CCXT (with graceful fallback if unavailable)
- Content creation automation via Selenium
- Income monitoring and reinvestment strategies
- Data analysis with Pandas/NumPy
- Excel reporting via OpenPyXL

Dependencies:
    pip install requests pandas numpy selenium openpyxl

Optional Dependencies:
    pip install ccxt  # For full trading functionality

Note: This is a demo/educational module. All external integrations are placeholders.
For production use, add proper API keys and security measures.

Security Notice:
- No hardcoded secrets or API keys
- Uses demo_mode=True by default to prevent unintended live trading
- Input validation on all public methods
- Secure logging (no sensitive data exposure)
"""

import logging
from typing import Dict, List, Any, Tuple, Union
import pandas as pd
import openpyxl
from openpyxl import Workbook
import os
import time
import random

# Attempt to import optional dependencies
try:
    import ccxt

    CCXT_AVAILABLE = True
except ImportError:
    CCXT_AVAILABLE = False
    ccxt = None  # type: ignore

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        WebDriverException,
        TimeoutException,
        NoSuchElementException,
    )

    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    webdriver = None  # type: ignore
    Options = None  # type: ignore


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("abundance_generator.log", mode="a"),
    ],
)
logger = logging.getLogger(__name__)


class InputValidationError(Exception):
    """Custom exception for input validation failures."""

    pass


class APIError(Exception):
    """Custom exception for API-related errors."""

    pass


class DependencyError(Exception):
    """Custom exception for missing optional dependencies."""

    pass


def validate_input(
    data: Any, expected_type: Union[type, Tuple[type, ...]], name: str = "data"
) -> bool:
    """
    Validate input data against expected type with additional checks.

    Args:
        data: Data to validate
        expected_type: Expected Python type or tuple of types
        name: Name of the parameter for error messages

    Returns:
        bool: True if valid

    Raises:
        InputValidationError: If validation fails
    """
    if not isinstance(data, expected_type):
        raise InputValidationError(
            f"Invalid type for {name}: expected {expected_type}, "
            f"got {type(data).__name__}"
        )

    # Additional validations based on type
    if isinstance(data, str):
        if not data.strip():
            raise InputValidationError(f"{name} cannot be empty string")
    elif isinstance(expected_type, tuple) and (
        int in expected_type or float in expected_type
    ):
        if not isinstance(data, (int, float)):
            raise InputValidationError(f"{name} must be numeric")

    return True


def log_error(message: str, level: str = "ERROR") -> None:
    """
    Log error messages with appropriate level.

    Args:
        message: Error message to log
        level: Log level (ERROR, WARNING, INFO)
    """
    if level == "ERROR":
        logger.error(message)
    elif level == "WARNING":
        logger.warning(message)
    else:
        logger.info(message)


class TradingEngine:
    """
    Handles algorithmic trading operations using CCXT.

    Supports multiple cryptocurrency exchanges via CCXT library.
    Placeholders for API credentials to maintain security.

    If CCXT is not available, operates in enhanced demo mode with simulated data.
    """

    def __init__(self, exchange_name: str = "binance", demo_mode: bool = True):
        """
        Initialize TradingEngine with exchange configuration.

        Args:
            exchange_name: Name of the exchange (e.g., 'binance', 'coinbase')
            demo_mode: If True, uses simulated API calls (no real trading)
        """
        self.exchange_name = exchange_name
        self.demo_mode = demo_mode
        self.exchange = None
        self.strategy_config = {}

        # If CCXT is not available, force demo mode and log warning
        if not CCXT_AVAILABLE and not demo_mode:
            logger.warning(
                "CCXT library not found. TradingEngine will operate in DEMO MODE only."
            )
            self.demo_mode = True

        # In demo mode, create a mock exchange instance
        if self.demo_mode:
            logger.info("TradingEngine initialized in DEMO MODE - no real trading")
            self._init_mock_exchange()
        elif CCXT_AVAILABLE:
            logger.info(f"TradingEngine initialized for {exchange_name}")
            # Note: Real exchange initialization happens in initialize_trading_strategy
        else:
            # This shouldn't happen due to the logic above, but as a fallback
            logger.warning("TradingEngine initialized with no CCXT and demo_mode=False")
            self.demo_mode = True
            self._init_mock_exchange()

    def _init_mock_exchange(self) -> None:
        """Initialize a mock exchange for demo/testing purposes."""

        class MockExchange:
            def __init__(self):
                self.name = "mock"
                self.has = {
                    "fetchOHLCV": True,
                    "fetchTicker": True,
                    "createOrder": True,
                    "fetchBalance": True,
                    "fetchOrderBook": True,
                }

            def fetch_ohlcv(self, symbol, timeframe, since=None, limit=100):
                """Mock OHLCV data for testing."""
                base_price = 100.0
                data = []
                for i in range(limit or 100):
                    base_price *= random.uniform(0.98, 1.02)
                    data.append(
                        [
                            i * 3600000,  # timestamp
                            base_price * random.uniform(0.99, 1.01),  # open
                            base_price * random.uniform(0.99, 1.01),  # high
                            base_price * random.uniform(0.99, 1.01),  # low
                            base_price * random.uniform(0.99, 1.01),  # close
                            base_price * random.uniform(10, 1000),  # volume
                        ]
                    )
                return data

            def fetch_ticker(self, symbol):
                """Mock ticker data."""
                return {
                    "symbol": symbol,
                    "last": 100.0,
                    "change": 0.5,
                    "percentage": 0.5,
                }

            def fetch_balance(self):
                """Mock balance data."""
                return {"info": {}, "total": {"USDT": 1000.0, "BTC": 0.01}}

            def fetch_order_book(self, symbol, limit=None):
                """Mock order book data."""
                return {
                    "symbol": symbol,
                    "bids": [[100.0, 1.0], [99.5, 2.0]],
                    "asks": [[100.5, 1.0], [101.0, 2.0]],
                }

        self.exchange = MockExchange()
        logger.info("Mock exchange initialized for safe testing")

    def _validate_trading_config(self, config: Dict) -> None:
        """
        Decomposed helper: Validates configuration dictionary structure and content.
        """
        validate_input(config, dict, "config")

        required_keys = ["symbol", "timeframe"]
        # Only require 'exchange' if not in demo mode (CCXT may be used)
        if not self.demo_mode:
            required_keys.append("exchange")

        missing_keys = [key for key in required_keys if key not in config]

        if missing_keys:
            raise InputValidationError(f"Missing required config keys: {missing_keys}")

        if not isinstance(config["symbol"], str) or not config["symbol"]:
            raise InputValidationError("Symbol must be a non-empty string")

    def _setup_ccxt_exchange(self, exchange_name: str) -> None:
        """
        Decomposed helper: Initializes the real CCXT exchange instance.
        """
        if not CCXT_AVAILABLE:
            raise DependencyError(
                "CCXT library is required for real exchange operations. "
                "Install with 'pip install ccxt' or enable demo_mode=True."
            )

        try:
            self.exchange = getattr(ccxt, exchange_name)()
            logger.info(f"CCXT exchange '{exchange_name}' initialized")
        except AttributeError:
            # Filter dir(ccxt) to show only public classes (exchanges)
            valid_exchanges = [
                name
                for name in dir(ccxt)
                if not name.startswith("_") and name[0].isupper()
            ]
            raise APIError(
                f"Exchange '{exchange_name}' not found. Valid exchanges: {valid_exchanges[:10]}..."
            )
        except Exception as e:
            logger.error(f"Failed to initialize exchange: {e}")
            raise APIError(f"Exchange initialization failed: {e}")

    def _setup_strategy_details(self, config: Dict) -> None:
        """
        Decomposed helper: Validates and logs strategy specifics like indicators and risk.
        """
        if "indicators" in config:
            valid_indicators = {"SMA", "EMA", "RSI", "MACD"}
            for indicator in config["indicators"]:
                if indicator not in valid_indicators:
                    logger.warning(f"Unknown indicator: {indicator}")

        if "risk_per_trade" in config:
            risk = config["risk_per_trade"]
            if not isinstance(risk, (int, float)) or risk <= 0 or risk > 0.1:
                logger.warning(
                    f"Risk per trade {risk} seems unreasonable, consider 0.01-0.05"
                )

    def initialize_trading_strategy(self, config: Dict) -> None:
        """
        Initialize trading strategy with configuration parameters.

        Args:
            config: Dictionary containing strategy configuration
                Required keys: 'symbol', 'timeframe'
                Optional keys: 'exchange', 'indicators', 'risk_per_trade'

        Example:
            config = {
                'exchange': 'binance',
                'symbol': 'BTC/USDT',
                'timeframe': '1h',
                'indicators': ['SMA', 'EMA'],
                'risk_per_trade': 0.02
            }

        Raises:
            InputValidationError: If config is invalid
            DependencyError: If CCXT is required but not available
        """
        try:
            # 1. Validate Structure
            self._validate_trading_config(config)

            # 2. Store Configuration
            self.strategy_config = config.copy()

            # 3. Initialize Exchange
            if not self.demo_mode:
                exchange_name = config.get("exchange", self.exchange_name)
                self._setup_ccxt_exchange(exchange_name)
            else:
                logger.info(f"Demo strategy configured for {config['symbol']}")

            # 4. Validate Details
            self._setup_strategy_details(config)

            logger.info(
                f"Trading strategy initialized successfully: {config['symbol']}"
            )

        except InputValidationError as e:
            log_error(f"Strategy validation failed: {e}")
            raise
        except Exception as e:
            log_error(f"Unexpected error in initialize_trading_strategy: {e}")
            raise APIError(f"Failed to initialize trading strategy: {e}")

    def _calculate_technical_indicators(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Decomposed helper: Calculates technical indicators (SMA, EMA, RSI, MACD).
        """
        # Simple Moving Average (SMA)
        if len(df) >= 10:
            df["sma_10"] = df["price"].rolling(window=10, min_periods=1).mean()
            df["sma_20"] = df["price"].rolling(window=20, min_periods=1).mean()

        # Exponential Moving Average (EMA)
        if len(df) >= 12:
            df["ema_12"] = df["price"].ewm(span=12, adjust=False).mean()
            df["ema_26"] = df["price"].ewm(span=26, adjust=False).mean()

        # RSI (Relative Strength Index)
        if len(df) >= 14:
            delta = df["price"].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()

            # Handle division by zero by checking if loss is zero or NaN
            # Using np.where to avoid ZeroDivisionError
            rs = gain / loss.replace(
                0, 1e-10
            )  # Replace 0 with small epsilon to avoid division by zero
            df["rsi"] = 100 - (100 / (1 + rs))

            # Cap RSI at 100 and handle NaN values
            df["rsi"] = df["rsi"].fillna(50)  # Neutral RSI for NaN cases
            df["rsi"] = df["rsi"].clip(0, 100)

        # MACD
        if len(df) >= 26:
            df["macd"] = df["ema_12"] - df["ema_26"]
            df["macd_signal"] = df["macd"].ewm(span=9, adjust=False).mean()
            df["macd_histogram"] = df["macd"] - df["macd_signal"]

        return df

    def _classify_trend(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Decomposed helper: Classifies market trend based on price changes.
        """
        if len(df) > 1:
            # Ensure price_change exists (calculated in analyze_market_trends)
            if "price_change" in df.columns:
                recent_changes = df["price_change"].tail(10).dropna()
                if len(recent_changes) > 0:
                    avg_change = recent_changes.mean()
                    if avg_change > 2:
                        df["trend"] = "Strong Bullish"
                    elif avg_change > 0.5:
                        df["trend"] = "Bullish"
                    elif avg_change < -2:
                        df["trend"] = "Strong Bearish"
                    elif avg_change < -0.5:
                        df["trend"] = "Bearish"
                    else:
                        df["trend"] = "Neutral"
        return df

    def analyze_market_trends(self, data: List[Dict]) -> pd.DataFrame:
        """
        Analyze market trends using Pandas/NumPy.

        Args:
            data: List of market data dictionaries. Each dict should contain:
                - 'timestamp': Unix timestamp
                - 'price': Current price
                - 'volume': Trading volume
                - Optional: 'open', 'high', 'low', 'close'

        Returns:
            pd.DataFrame: Analyzed market data with trend indicators

        Raises:
            InputValidationError: If data is invalid
        """
        try:
            # Validate input
            validate_input(data, list, "data")

            if not data:
                raise InputValidationError("Data list cannot be empty")

            # Convert to DataFrame
            df = pd.DataFrame(data)

            # Validate required columns AFTER creating DataFrame
            required_cols = ["timestamp", "price", "volume"]
            missing_cols = [col for col in required_cols if col not in df.columns]

            if missing_cols:
                raise InputValidationError(
                    f"Missing required data columns: {missing_cols}"
                )

            # Sort by timestamp
            df = df.sort_values("timestamp").reset_index(drop=True)

            # 1. Calculate technical indicators (Decomposed)
            df = self._calculate_technical_indicators(df)

            # Price changes and trends
            # Calculate price_change BEFORE calling _classify_trend
            df["price_change"] = df["price"].pct_change() * 100
            df["volume_change"] = df["volume"].pct_change() * 100

            # 2. Classify trend (Decomposed)
            df = self._classify_trend(df)

            # Add summary statistics
            df["max_price"] = df["price"].max()
            df["min_price"] = df["price"].min()
            df["avg_price"] = df["price"].mean()

            logger.info(f"Market analysis complete: {len(df)} data points processed")
            return df

        except InputValidationError as e:
            log_error(f"Market trend analysis validation failed: {e}")
            raise
        except Exception as e:
            log_error(f"Unexpected error in analyze_market_trends: {e}")
            raise APIError(f"Failed to analyze market trends: {e}")


class ContentManager:
    """
    Handles content creation and monetization automation.

    Uses Selenium for browser automation. Requires ChromeDriver.
    In demo mode, simulates actions without actual browser interaction.
    """

    def __init__(self, demo_mode: bool = True):
        """
        Initialize ContentManager.

        Args:
            demo_mode: If True, simulates browser actions (no real browser)
        """
        self.demo_mode = demo_mode
        self.driver = None
        self.setup_complete = False

        # Check if Selenium is available
        if not SELENIUM_AVAILABLE and not demo_mode:
            logger.warning(
                "Selenium library not found. ContentManager will operate in DEMO MODE only."
            )
            self.demo_mode = True

        if self.demo_mode:
            logger.info("ContentManager initialized in DEMO MODE - simulated actions")
        else:
            logger.info("ContentManager initialized - browser automation ready")

    def _setup_driver(self) -> None:
        """Setup Chrome WebDriver for browser automation."""
        if not SELENIUM_AVAILABLE:
            raise DependencyError(
                "Selenium library is required for browser automation. "
                "Install with 'pip install selenium' or enable demo_mode=True."
            )

        try:
            # FIXED: Ensure Options is instantiated correctly
            # Importing Options as a class and instantiating it with ()
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # Run in headless mode
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")

            self.driver = webdriver.Chrome(options=chrome_options)
            self.setup_complete = True
            logger.info("Chrome WebDriver initialized successfully")

        except WebDriverException as e:
            logger.error(f"Failed to initialize WebDriver: {e}")
            raise APIError(f"WebDriver setup failed: {e}")

    def close(self) -> None:
        """
        Properly close the WebDriver instance to prevent zombie processes.

        This method should be called when the ContentManager is no longer needed.
        """
        if self.driver is not None:
            try:
                self.driver.quit()
                self.driver = None
                self.setup_complete = False
                logger.info("WebDriver closed successfully")
            except Exception as e:
                logger.error(f"Error closing WebDriver: {e}")

    def automate_content_creation(self, topic: str, platform: str) -> str:
        """
        Automate content creation and posting.

        Args:
            topic: Content topic or title
            platform: Target platform (e.g., 'twitter', 'medium', 'linkedin')

        Returns:
            str: Success message with content ID

        Raises:
            InputValidationError: If inputs are invalid
            DependencyError: If Selenium is required but not available
        """
        try:
            # Validate inputs
            validate_input(topic, str, "topic")
            validate_input(platform, str, "platform")

            # Sanitize platform
            platform = platform.lower().strip()
            valid_platforms = ["twitter", "medium", "linkedin", "reddit"]
            if platform not in valid_platforms:
                logger.warning(
                    f"Platform '{platform}' is not in known list. Proceeding with caution."
                )

            if self.demo_mode:
                # Simulate content creation
                content_id = f"{platform}_{int(time.time())}"
                logger.info(
                    f"[DEMO] Created content on {platform}: '{topic}' (ID: {content_id})"
                )

                # Simulate delay
                time.sleep(0.1)

                return f"Content '{topic}' successfully posted to {platform} with ID: {content_id}"

            # Real browser automation
            if not self.setup_complete:
                self._setup_driver()

            # Simulate browser actions for demonstration
            # Note: Actual implementation would require specific platform logic
            try:
                # This is a placeholder - actual implementation would navigate to platform
                # and fill forms, click buttons, etc.
                logger.info(f"Simulating content creation on {platform}...")

                # Example: Mock a visit and form submission
                time.sleep(0.5)  # Simulate page load

                content_id = f"{platform}_{int(time.time())}"
                success_message = f"Content '{topic}' successfully posted to {platform} (ID: {content_id})"

                logger.info(success_message)
                return success_message

            except (TimeoutException, NoSuchElementException) as e:
                logger.error(f"Browser automation failed: {e}")
                raise APIError(f"Failed to create content on {platform}: {e}")

        except InputValidationError as e:
            log_error(f"Content creation validation failed: {e}")
            raise
        except Exception as e:
            log_error(f"Unexpected error in automate_content_creation: {e}")
            raise APIError(f"Failed to automate content creation: {e}")


class AssetMonitor:
    """
    Handles monitoring of income streams and reinvestment decisions.

    Tracks profits from various sources and manages Excel reporting.
    """

    # Define strategy allocation constants to avoid hardcoded magic numbers
    STRATEGY_ALLOCATIONS = {
        "trading": {
            "trading_account": 0.8,
            "stable_reserve": 0.15,
            "emergency_fund": 0.05,
        },
        "content": {"content_creation": 0.6, "marketing": 0.3, "tools": 0.1},
        "balanced": {
            "trading_account": 0.4,
            "content_creation": 0.3,
            "stable_reserve": 0.2,
            "emergency_fund": 0.1,
        },
        "conservative": {
            "stable_reserve": 0.5,
            "emergency_fund": 0.3,
            "trading_account": 0.15,
            "content_creation": 0.05,
        },
    }

    def __init__(self, excel_file: str = "income_tracker.xlsx"):
        """
        Initialize AssetMonitor.

        Args:
            excel_file: Path to Excel file for tracking income
        """
        self.excel_file = excel_file
        self.income_streams = {}
        logger.info(f"AssetMonitor initialized with Excel file: {excel_file}")

    def monitor_income_streams(self, streams: Dict) -> Dict:
        """
        Monitor and update income streams.

        Args:
            streams: Dictionary of income streams with current values
                Example: {'trading': 150.50, 'content': 25.00, 'other': 10.00}

        Returns:
            Dict: Updated income streams with calculations

        Raises:
            InputValidationError: If streams is invalid
        """
        try:
            # Validate input
            validate_input(streams, dict, "streams")

            if not streams:
                raise InputValidationError("Streams dictionary cannot be empty")

            # Validate stream values
            for stream_name, amount in streams.items():
                if not isinstance(amount, (int, float)):
                    raise InputValidationError(
                        f"Invalid amount for '{stream_name}': must be numeric"
                    )
                if amount < 0:
                    raise InputValidationError(
                        f"Invalid amount for '{stream_name}': cannot be negative"
                    )

            # Update internal tracking
            self.income_streams.update(streams)

            # Calculate totals
            total_income = sum(streams.values())
            stream_count = len(streams)

            # Create summary
            summary = {
                "streams": streams.copy(),
                "total_income": total_income,
                "stream_count": stream_count,
                "average_per_stream": total_income / stream_count
                if stream_count > 0
                else 0,
                "timestamp": time.time(),
            }

            # Log to Excel
            log_success = self._log_to_excel(streams, total_income)
            if not log_success:
                summary["excel_log_error"] = True
                logger.warning("Excel logging failed but monitoring continued")

            logger.info(
                f"Income monitoring complete: {stream_count} streams, ${total_income:.2f} total"
            )
            return summary

        except InputValidationError as e:
            log_error(f"Income monitoring validation failed: {e}")
            raise
        except Exception as e:
            log_error(f"Unexpected error in monitor_income_streams: {e}")
            raise APIError(f"Failed to monitor income streams: {e}")

    def _log_to_excel(self, streams: Dict, total_income: float) -> bool:
        """
        Log income data to Excel file in tall format (timestamp, stream_name, amount).

        FIXED: Uses tall format instead of wide to prevent column overlap.
        Returns success status instead of swallowing exceptions.

        Args:
            streams: Dictionary of income streams
            total_income: Total income amount

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Check if file exists
            file_exists = os.path.exists(self.excel_file)

            if file_exists:
                # Load existing workbook
                try:
                    wb = openpyxl.load_workbook(self.excel_file)
                    ws = wb.active
                except Exception as e:
                    logger.warning(
                        f"Failed to load existing Excel file: {e}. Creating new one."
                    )
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Income Tracker"
                    self._write_headers(ws)
            else:
                # Create new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Income Tracker"
                self._write_headers(ws)

            # Find next empty row
            row = ws.max_row + 1

            # FIXED: Write each stream as a separate row (tall format)
            # Format: Timestamp | Stream Name | Amount
            timestamp_str = time.strftime("%Y-%m-%d %H:%M:%S")

            for stream_name, amount in streams.items():
                ws.cell(row=row, column=1, value=timestamp_str)
                ws.cell(row=row, column=2, value=stream_name)
                ws.cell(row=row, column=3, value=amount)
                row += 1

            # Save workbook
            wb.save(self.excel_file)
            logger.debug(f"Income data logged to {self.excel_file}")
            return True

        except Exception as e:
            logger.error(f"Failed to write to Excel file: {e}")
            return False

    def _write_headers(self, ws) -> None:
        """Write headers to Excel worksheet (tall format)."""
        headers = ["Timestamp", "Stream Name", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

    def reinvest_profits(self, profits: float, strategy: str) -> bool:
        """
        Reinvest profits according to strategy.

        Args:
            profits: Total profit amount to reinvest
            strategy: Reinvestment strategy ('trading', 'content', 'balanced', 'conservative')

        Returns:
            bool: True if reinvestment was successful

        Raises:
            InputValidationError: If inputs are invalid
        """
        try:
            # Validate inputs
            validate_input(profits, (int, float), "profits")
            validate_input(strategy, str, "strategy")

            if profits <= 0:
                raise InputValidationError("Profits must be greater than zero")

            # Sanitize strategy
            strategy = strategy.lower().strip()
            valid_strategies = list(self.STRATEGY_ALLOCATIONS.keys())

            if strategy not in valid_strategies:
                logger.warning(
                    f"Strategy '{strategy}' is not standard. Using 'balanced'."
                )
                strategy = "balanced"

            # Calculate allocations
            allocations = self._calculate_allocations(profits, strategy)

            # Simulate reinvestment process
            logger.info(
                f"Starting reinvestment of ${profits:.2f} using '{strategy}' strategy"
            )

            # Validate allocations
            total_allocated = sum(allocations.values())
            if abs(total_allocated - profits) > 0.01:
                logger.error(f"Allocation mismatch: {total_allocated} != {profits}")
                return False

            # Log allocations
            for asset, amount in allocations.items():
                if amount > 0:
                    logger.info(f"Allocating ${amount:.2f} to {asset}")

            # Simulate transaction delay
            time.sleep(0.1)

            # Update internal tracking
            self.income_streams["reinvested"] = profits

            logger.info(
                f"Reinvestment complete: ${profits:.2f} allocated via '{strategy}'"
            )
            return True

        except InputValidationError as e:
            log_error(f"Reinvestment validation failed: {e}")
            return False
        except Exception as e:
            log_error(f"Unexpected error in reinvest_profits: {e}")
            return False

    def _calculate_allocations(self, profits: float, strategy: str) -> Dict[str, float]:
        """
        Calculate allocation amounts based on strategy.

        Uses class constant STRATEGY_ALLOCATIONS instead of hardcoded values.

        Args:
            profits: Total profit amount
            strategy: Strategy name

        Returns:
            Dict: Asset -> amount mapping
        """
        allocations = {}

        if strategy in self.STRATEGY_ALLOCATIONS:
            ratios = self.STRATEGY_ALLOCATIONS[strategy]
            for asset, ratio in ratios.items():
                allocations[asset] = profits * ratio
        else:
            # Fallback to balanced if strategy not found
            ratios = self.STRATEGY_ALLOCATIONS["balanced"]
            for asset, ratio in ratios.items():
                allocations[asset] = profits * ratio

        return allocations


class AbundanceGenerator:
    """
    Main orchestrator class for the AbundanceGenerator system.

    Integrates TradingEngine, ContentManager, and AssetMonitor to provide
    a unified interface for automated resource acquisition.

    Implements context manager protocol for proper resource cleanup.
    """

    def __init__(
        self, exchange: str = "binance", excel_file: str = "income_tracker.xlsx"
    ):
        """
        Initialize AbundanceGenerator with all components.

        Args:
            exchange: Trading exchange name (default: 'binance')
            excel_file: Path to Excel tracking file (default: 'income_tracker.xlsx')
        """
        self.logger = logging.getLogger(__name__)
        self.logger.info("Initializing AbundanceGenerator...")

        # Initialize components in demo mode by default for safety
        self.trading = TradingEngine(exchange, demo_mode=True)
        self.content = ContentManager(demo_mode=True)
        self.monitor = AssetMonitor(excel_file)

        self.logger.info("AbundanceGenerator initialized successfully")

    def __enter__(self):
        """Enter context manager."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Exit context manager with resource cleanup."""
        self.cleanup()
        return False

    def cleanup(self) -> None:
        """Clean up resources (e.g., WebDriver)."""
        if hasattr(self, "content"):
            self.content.close()
            self.logger.info("AbundanceGenerator resources cleaned up")

    def initialize_trading_strategy(self, config: Dict) -> None:
        """
        Initialize trading strategy with configuration.

        This is the specified entry point method.

        Args:
            config: Trading configuration dictionary

        Example:
            generator = AbundanceGenerator()
            config = {
                'exchange': 'binance',
                'symbol': 'BTC/USDT',
                'timeframe': '1h',
                'indicators': ['SMA', 'EMA', 'RSI']
            }
            generator.initialize_trading_strategy(config)

        Raises:
            InputValidationError: If config is invalid
            APIError: If trading initialization fails
            DependencyError: If CCXT is required but not available
        """
        self.trading.initialize_trading_strategy(config)

    def analyze_market_trends(self, data: List[Dict]) -> pd.DataFrame:
        """
        Analyze market trends with Pandas/NumPy.

        This is the specified entry point method.

        Args:
            data: List of market data dictionaries

        Returns:
            pd.DataFrame: Analyzed data with trends

        Example:
            market_data = [
                {'timestamp': 1234567890, 'price': 45000, 'volume': 1000},
                {'timestamp': 1234567900, 'price': 45200, 'volume': 1100}
            ]
            df = generator.analyze_market_trends(market_data)
            print(df[['price', 'sma_10', 'rsi', 'trend']])

        Raises:
            InputValidationError: If data is invalid
            APIError: If analysis fails
        """
        return self.trading.analyze_market_trends(data)

    def automate_content_creation(self, topic: str, platform: str) -> str:
        """
        Automate content creation and posting.

        This is the specified entry point method.

        Args:
            topic: Content topic or title
            platform: Target platform (e.g., 'twitter', 'medium', 'linkedin')

        Returns:
            str: Success message with content ID

        Example:
            result = generator.automate_content_creation(
                topic="My Trading Strategy Tips",
                platform="twitter"
            )
            print(result)

        Raises:
            InputValidationError: If inputs are invalid
            APIError: If content creation fails
            DependencyError: If Selenium is required but not available
        """
        return self.content.automate_content_creation(topic, platform)

    def monitor_income_streams(self, streams: Dict) -> Dict:
        """
        Monitor and track income from various streams.

        This is the specified entry point method.

        Args:
            streams: Dictionary of income streams with amounts
                Example: {'trading': 150.50, 'content': 25.00, 'affiliate': 10.00}

        Returns:
            Dict: Summary with total and statistics

        Example:
            income = {'trading': 250.00, 'content': 75.50}
            summary = generator.monitor_income_streams(income)
            print(f"Total: ${summary['total_income']:.2f}")

        Raises:
            InputValidationError: If streams is invalid
            APIError: If monitoring fails
        """
        return self.monitor.monitor_income_streams(streams)

    def reinvest_profits(self, profits: float, strategy: str) -> bool:
        """
        Reinvest profits according to strategy.

        This is the specified entry point method.

        Args:
            profits: Total profit amount
            strategy: Reinvestment strategy ('trading', 'content', 'balanced', 'conservative')

        Returns:
            bool: True if successful

        Example:
            success = generator.reinvest_profits(1000.00, 'balanced')
            if success:
                print("Reinvestment completed")
            else:
                print("Reinvestment failed")

        Raises:
            InputValidationError: If inputs are invalid
        """
        return self.monitor.reinvest_profits(profits, strategy)

    def get_system_status(self) -> Dict:
        """
        Get current status of all components.

        Returns:
            Dict: System status information
        """
        status = {
            "trading_engine": {
                "exchange": self.trading.exchange_name,
                "demo_mode": self.trading.demo_mode,
                "strategy_configured": bool(self.trading.strategy_config),
                "ccxt_available": CCXT_AVAILABLE,
            },
            "content_manager": {
                "demo_mode": self.content.demo_mode,
                "setup_complete": self.content.setup_complete,
                "selenium_available": SELENIUM_AVAILABLE,
            },
            "asset_monitor": {
                "excel_file": self.monitor.excel_file,
                "streams_tracked": len(self.monitor.income_streams),
            },
        }
        return status


# Module-level helper functions for convenience
def generate_report(generator: AbundanceGenerator) -> str:
    """
    Generate a summary report of current system status.

    Args:
        generator: Initialized AbundanceGenerator instance

    Returns:
        str: Formatted report string
    """
    try:
        status = generator.get_system_status()

        report = []
        report.append("=== ABUNDANCE GENERATOR SYSTEM REPORT ===")
        report.append(f"Timestamp: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")

        # Trading status
        t = status["trading_engine"]
        report.append("TRADING ENGINE:")
        report.append(f"  Exchange: {t['exchange']}")
        report.append(f"  Mode: {'DEMO' if t['demo_mode'] else 'LIVE'}")
        report.append(
            f"  Strategy: {'Configured' if t['strategy_configured'] else 'Not Configured'}"
        )
        report.append(f"  CCXT Available: {'Yes' if t['ccxt_available'] else 'No'}")
        report.append("")

        # Content status
        c = status["content_manager"]
        report.append("CONTENT MANAGER:")
        report.append(f"  Mode: {'DEMO' if c['demo_mode'] else 'LIVE'}")
        report.append(
            f"  Setup: {'Complete' if c['setup_complete'] else 'Not Complete'}"
        )
        report.append(
            f"  Selenium Available: {'Yes' if c['selenium_available'] else 'No'}"
        )
        report.append("")

        # Asset monitor status
        m = status["asset_monitor"]
        report.append("ASSET MONITOR:")
        report.append(f"  Excel File: {m['excel_file']}")
        report.append(f"  Streams Tracked: {m['streams_tracked']}")
        report.append("")

        return "\n".join(report)

    except Exception as e:
        logger.error(f"Failed to generate report: {e}")
        return f"Error generating report: {e}"


# Example usage function
def example_usage():
    """
    Example function demonstrating how to use the AbundanceGenerator.

    This function shows basic usage patterns and can be run as a script
    to test the module functionality.
    """
    print("=== AbundanceGenerator Demo ===")
    print()

    try:
        # Initialize the generator (using context manager for proper cleanup)
        print("1. Initializing AbundanceGenerator...")
        with AbundanceGenerator() as generator:
            # Example 1: Initialize trading strategy
            print("\n2. Initializing trading strategy...")
            config = {
                "exchange": "binance",
                "symbol": "BTC/USDT",
                "timeframe": "1h",
                "indicators": ["SMA", "EMA", "RSI"],
                "risk_per_trade": 0.02,
            }
            generator.initialize_trading_strategy(config)
            print("   ✓ Trading strategy initialized")

            # Example 2: Analyze market trends
            print("\n3. Analyzing market trends...")
            sample_data = [
                {
                    "timestamp": 1672531200 + i * 3600,
                    "price": 45000 + i * 10,
                    "volume": 1000 + i * 50,
                }
                for i in range(20)
            ]
            df = generator.analyze_market_trends(sample_data)
            print(f"   ✓ Analyzed {len(df)} data points")
            print(
                f"   Sample trend: {df['trend'].iloc[-1] if 'trend' in df.columns else 'N/A'}"
            )

            # Example 3: Automate content creation
            print("\n4. Creating content...")
            content_result = generator.automate_content_creation(
                topic="Trading Strategy Insights", platform="twitter"
            )
            print(f"   ✓ {content_result}")

            # Example 4: Monitor income streams
            print("\n5. Monitoring income streams...")
            income = {"trading": 250.50, "content": 75.00, "affiliate": 12.50}
            summary = generator.monitor_income_streams(income)
            print(f"   ✓ Total income: ${summary['total_income']:.2f}")
            print(f"   ✓ Streams tracked: {summary['stream_count']}")

            # Example 5: Reinvest profits
            print("\n6. Reinvesting profits...")
            success = generator.reinvest_profits(500.00, "balanced")
            if success:
                print("   ✓ Reinvestment completed")
            else:
                print("   ✗ Reinvestment failed")

            # Example 6: Get system status
            print("\n7. System Status:")
            print(generator.get_system_status())

            # Generate report
            print("\n8. Final Report:")
            print(generate_report(generator))

        print("\n=== Demo completed successfully ===")

    except Exception as e:
        print(f"\n✗ Demo failed with error: {e}")
        logger.error(f"Demo execution error: {e}")


# Run example if script is executed directly
if __name__ == "__main__":
    example_usage()
